from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum, Count
from datetime import date, timedelta, datetime
import json, random

from .models import Bus, Book, UserProfile, PromoCode, Review, FavouriteRoute, WalletTransaction, Operator

import qrcode
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.pagesizes import A4

# ── All Indian Cities ──────────────────────────────────────────────────────────
CITIES = [
    # Maharashtra
    "Mumbai","Pune","Nagpur","Nashik","Aurangabad","Solapur","Kolhapur","Thane",
    "Navi Mumbai","Amravati","Akola","Latur","Nanded","Satara","Sangli","Jalgaon",
    "Ahmednagar","Chandrapur","Dhule","Ratnagiri","Sindhudurg","Wardha","Yavatmal",
    # Delhi & NCR
    "Delhi","New Delhi","Noida","Gurgaon","Faridabad","Ghaziabad","Greater Noida",
    # Uttar Pradesh
    "Lucknow","Agra","Varanasi","Kanpur","Allahabad","Prayagraj","Meerut","Bareilly",
    "Aligarh","Moradabad","Gorakhpur","Mathura","Vrindavan","Jhansi","Firozabad",
    # Rajasthan
    "Jaipur","Jodhpur","Udaipur","Kota","Ajmer","Bikaner","Alwar","Pushkar",
    "Mount Abu","Chittorgarh","Jaisalmer","Sikar","Bharatpur",
    # Gujarat
    "Ahmedabad","Surat","Vadodara","Rajkot","Bhavnagar","Jamnagar","Gandhinagar",
    "Anand","Junagadh","Nadiad","Bharuch","Mehsana","Porbandar","Dwarka",
    # Karnataka
    "Bangalore","Mysore","Hubli","Dharwad","Mangalore","Belgaum","Gulbarga",
    "Davangere","Bijapur","Shimoga","Tumkur","Hassan","Udupi","Hampi",
    # Tamil Nadu
    "Chennai","Coimbatore","Madurai","Tiruchirappalli","Salem","Tirunelveli",
    "Vellore","Erode","Thoothukudi","Thanjavur","Kanyakumari","Ooty","Kodaikanal",
    # Andhra Pradesh & Telangana
    "Hyderabad","Secunderabad","Visakhapatnam","Vijayawada","Guntur","Tirupati",
    "Nellore","Kurnool","Rajahmundry","Kakinada","Warangal","Nizamabad","Karimnagar",
    # Kerala
    "Thiruvananthapuram","Kochi","Kozhikode","Thrissur","Kollam","Alappuzha",
    "Palakkad","Kannur","Kasaragod","Munnar","Wayanad","Varkala","Kottayam",
    # West Bengal
    "Kolkata","Howrah","Asansol","Siliguri","Durgapur","Bardhaman","Darjeeling",
    "Kalimpong","Malda","Haldia","Kharagpur","Midnapore",
    # Madhya Pradesh
    "Bhopal","Indore","Gwalior","Jabalpur","Ujjain","Sagar","Dewas","Satna",
    "Ratlam","Rewa","Mandsaur","Khajuraho","Pachmarhi",
    # Punjab & Haryana
    "Chandigarh","Ludhiana","Amritsar","Jalandhar","Patiala","Bathinda",
    "Rohtak","Hisar","Panipat","Sonipat","Karnal","Ambala","Gurdaspur",
    # Bihar & Jharkhand
    "Patna","Gaya","Muzaffarpur","Bhagalpur","Darbhanga","Ranchi","Dhanbad",
    "Jamshedpur","Bokaro","Deoghar","Hazaribagh",
    # Odisha
    "Bhubaneswar","Cuttack","Rourkela","Berhampur","Puri","Sambalpur","Konark",
    # Assam & Northeast
    "Guwahati","Dibrugarh","Silchar","Jorhat","Tezpur","Shillong","Agartala",
    "Imphal","Kohima","Itanagar","Aizawl","Gangtok",
    # Himachal Pradesh & Uttarakhand
    "Shimla","Manali","Dharamshala","Mandi","Kullu","Dehradun","Haridwar",
    "Rishikesh","Nainital","Mussoorie","Almora","Roorkee",
    # Goa
    "Panaji","Margao","Vasco da Gama","Mapusa","Ponda",
    # Jammu & Kashmir
    "Jammu","Srinagar","Leh","Katra","Udhampur",
    # Chhattisgarh
    "Raipur","Bhilai","Bilaspur","Korba","Durg",
    # Other popular
    "Amritsar","Varanasi","Haridwar","Rishikesh","Agra","Mathura","Vrindavan",
    "Tirupati","Shirdi","Ajmer","Pushkar","Dwarka","Somnath",
]
CITIES = sorted(list(set(CITIES)))


# ── helpers ───────────────────────────────────────────────────────────────────
def get_or_create_profile(user):
    profile, _ = UserProfile.objects.get_or_create(user=user)
    return profile

def send_email_safe(subject, body, to):
    try:
        send_mail(subject, body, settings.EMAIL_HOST_USER, [to], fail_silently=True)
    except Exception:
        pass


# ── HOME ──────────────────────────────────────────────────────────────────────
def home(request):
    if request.user.is_authenticated:
        popular_routes = [
            {'emoji':'🏙️','from':'Mumbai','to':'Pune','price':250},
            {'emoji':'🌴','from':'Pune','to':'Goa','price':600},
            {'emoji':'🏰','from':'Delhi','to':'Jaipur','price':350},
            {'emoji':'🌊','from':'Chennai','to':'Bangalore','price':400},
            {'emoji':'🕌','from':'Hyderabad','to':'Mumbai','price':700},
            {'emoji':'🎭','from':'Bangalore','to':'Chennai','price':380},
            {'emoji':'🏔️','from':'Delhi','to':'Manali','price':900},
            {'emoji':'🎪','from':'Kolkata','to':'Puri','price':300},
        ]
        how_steps = [
            {'icon':'🔍','title':'Search Route','desc':'Enter source, destination and date','bg':'rgba(56,189,248,.15)'},
            {'icon':'💺','title':'Pick Seats','desc':'Choose your preferred seats visually','bg':'rgba(34,197,94,.15)'},
            {'icon':'💳','title':'Pay Securely','desc':'Use card, promo or wallet balance','bg':'rgba(251,191,36,.15)'},
            {'icon':'🎫','title':'Get Ticket','desc':'Download PDF with QR code instantly','bg':'rgba(167,139,250,.15)'},
        ]
        return render(request, 'myapp/home.html', {
            'popular_routes': popular_routes,
            'how_steps': how_steps,
        })
    return render(request, 'myapp/signin.html')


# ── FIND BUS ──────────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def findbus(request):
    context = {'cities': CITIES}
    fav_routes = FavouriteRoute.objects.filter(user=request.user)
    context['fav_routes'] = fav_routes

    if request.method == 'POST':
        source_r    = request.POST.get('source', '').strip()
        dest_r      = request.POST.get('destination', '').strip()
        date_r      = request.POST.get('date', '').strip()
        sort_by     = request.POST.get('sort', 'time')
        is_round    = request.POST.get('round_trip') == 'on'
        return_date = request.POST.get('return_date', '').strip()

        if not source_r or not dest_r or not date_r:
            context['error'] = "Please fill all fields."
            return render(request, 'myapp/findbus.html', context)
        if source_r.lower() == dest_r.lower():
            context['error'] = "Source and destination cannot be the same."
            return render(request, 'myapp/findbus.html', context)
        try:
            search_date = datetime.strptime(date_r, '%Y-%m-%d').date()
            if search_date < date.today():
                context['error'] = "Please select today or a future date."
                return render(request, 'myapp/findbus.html', context)
        except ValueError:
            context['error'] = "Invalid date."
            return render(request, 'myapp/findbus.html', context)

        bus_list = Bus.objects.filter(
            source__iexact=source_r, dest__iexact=dest_r,
            date=date_r, rem__gt=0
        )
        if sort_by == 'price':
            bus_list = bus_list.order_by('base_price')
        elif sort_by == 'seats':
            bus_list = bus_list.order_by('-rem')
        else:
            bus_list = bus_list.order_by('time')

        nearby = []
        if not bus_list.exists():
            nearby = list(Bus.objects.filter(
                dest__iexact=dest_r, date=date_r, rem__gt=0
            ).values_list('source', flat=True).distinct()[:5])

        return_buses = []
        if is_round and return_date:
            return_buses = Bus.objects.filter(
                source__iexact=dest_r, dest__iexact=source_r,
                date=return_date, rem__gt=0
            )

        return render(request, 'myapp/list.html', {
            'bus_list': bus_list,
            'source': source_r, 'dest': dest_r, 'date': date_r,
            'sort_by': sort_by, 'nearby': nearby,
            'is_round': is_round, 'return_buses': return_buses, 'return_date': return_date,
        })

    return render(request, 'myapp/findbus.html', context)


# ── SAVE / DELETE FAVOURITE ───────────────────────────────────────────────────
@login_required(login_url='signin')
def save_favourite(request):
    if request.method == 'POST':
        src = request.POST.get('source', '').strip()
        dst = request.POST.get('dest', '').strip()
        if src and dst:
            FavouriteRoute.objects.get_or_create(user=request.user, source=src, dest=dst)
            messages.success(request, f"Route {src} → {dst} saved!")
    return redirect('findbus')

@login_required(login_url='signin')
def delete_favourite(request, fav_id):
    fav = get_object_or_404(FavouriteRoute, id=fav_id, user=request.user)
    fav.delete()
    return redirect('findbus')


# ── SELECT SEATS ──────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def select_seats(request, bus_id):
    bus = get_object_or_404(Bus, id=bus_id)
    bookings = Book.objects.filter(busid=bus.id, status='B')
    booked_seats = []
    for b in bookings:
        if b.seat_numbers:
            for s in b.seat_numbers.split(","):
                s = s.strip()
                if s.isdigit():
                    booked_seats.append(int(s))

    dynamic_price = bus.dynamic_price()
    return render(request, 'myapp/select_seats.html', {
        'bus': bus,
        'booked_seats': booked_seats,
        'total_seats': range(1, bus.nos + 1),
        'dynamic_price': dynamic_price,
        'price_surge': dynamic_price > float(bus.base_price),
    })


# ── CONFIRM BOOKING ───────────────────────────────────────────────────────────
@login_required(login_url='signin')
def confirm_booking(request, bus_id):
    bus = get_object_or_404(Bus, id=bus_id)

    if request.method == "POST":

        # ── STEP 2: PAYMENT SUBMITTED ──────────────────────────────────────
        if request.POST.get("payment_done") == "yes":
            seat_string   = request.POST.get("seat_string", "")
            promo         = request.POST.get("promo_applied", "").strip()
            use_wallet    = request.POST.get("use_wallet") == "1"
            try:
                total_seats   = int(request.POST.get("total_seats", 0))
                total_cost    = float(request.POST.get("total_cost", 0))
                original_cost = float(request.POST.get("original_cost", total_cost))
                discount_amt  = float(request.POST.get("discount_amount", 0))
            except (ValueError, TypeError):
                return render(request, 'myapp/error.html', {'error': 'Invalid booking data. Please try again.'})

            if total_seats <= 0:
                return render(request, 'myapp/error.html', {'error': 'No seats selected.'})

            # Collect passenger details
            passenger_details = []
            for i in range(1, total_seats + 1):
                passenger_details.append({
                    'name':   request.POST.get(f'p_name_{i}', f'Passenger {i}'),
                    'age':    request.POST.get(f'p_age_{i}', ''),
                    'gender': request.POST.get(f'p_gender_{i}', ''),
                })

            profile = get_or_create_profile(request.user)
            wallet_used = 0.0

            with transaction.atomic():
                bus_locked = Bus.objects.select_for_update().get(id=bus_id)
                if bus_locked.rem < total_seats:
                    return render(request, 'myapp/error.html',
                                  {'error': 'Sorry, seats were just taken. Please re-select.'})

                # Double-booking guard
                already_booked = []
                for b in Book.objects.filter(busid=bus.id, status='B'):
                    if b.seat_numbers:
                        already_booked.extend([s.strip() for s in b.seat_numbers.split(",")])
                conflict = [s.strip() for s in seat_string.split(",") if s.strip() in already_booked]
                if conflict:
                    return render(request, 'myapp/error.html',
                                  {'error': f'Seats {", ".join(conflict)} already booked. Please re-select.'})

                # Wallet deduction
                if use_wallet and float(profile.wallet_balance) > 0:
                    wallet_used = min(float(profile.wallet_balance), total_cost)
                    total_cost  = round(total_cost - wallet_used, 2)
                    profile.wallet_balance = float(profile.wallet_balance) - wallet_used
                    profile.save()
                    WalletTransaction.objects.create(
                        user=request.user, amount=wallet_used,
                        transaction_type='debit',
                        description=f'Used for booking {bus.bus_name}'
                    )

                # Promo used
                if promo:
                    try:
                        pc = PromoCode.objects.get(code__iexact=promo)
                        if pc.is_valid():
                            pc.used_count += 1
                            pc.save()
                    except PromoCode.DoesNotExist:
                        pass

                book = Book.objects.create(
                    name=request.user.username,
                    email=request.user.email,
                    userid=request.user.id,
                    busid=bus.id,
                    bus_name=bus.bus_name,
                    source=bus.source,
                    dest=bus.dest,
                    nos=total_seats,
                    seat_numbers=seat_string,
                    original_price=original_cost,
                    price=total_cost,
                    promo_code=promo or None,
                    discount_amount=discount_amt,
                    date=bus.date,
                    time=bus.time,
                    status='B',
                    passenger_details=json.dumps(passenger_details),
                )
                bus_locked.rem -= total_seats
                bus_locked.save()

            send_email_safe(
                'RouteWay – Booking Confirmed!',
                f"""Hello {request.user.username},

Your booking is confirmed!

Booking ID : #{book.id}
Bus        : {bus.bus_name}
From       : {bus.source}
To         : {bus.dest}
Seat(s)    : {seat_string}
Date       : {bus.date}
Time       : {bus.time}
Total Fare : Rs.{total_cost}

Thank you for choosing RouteWay!
""",
                request.user.email
            )

            return render(request, 'myapp/bookings.html', {
                'book': book,
                'cost': total_cost,
                'wallet_used': wallet_used,
                'passenger_details': passenger_details,
            })

        # ── STEP 1: SEATS SELECTED ─────────────────────────────────────────
        selected_seats = request.POST.getlist('seats')
        if not selected_seats:
            messages.warning(request, "Please select at least one seat.")
            return redirect('select_seats', bus_id=bus.id)

        if bus.rem < len(selected_seats):
            return render(request, 'myapp/error.html', {'error': 'Not enough seats available.'})

        seat_string   = ",".join(selected_seats)
        total_seats   = len(selected_seats)
        dynamic_price = bus.dynamic_price()
        total_cost    = round(total_seats * dynamic_price, 2)
        profile       = get_or_create_profile(request.user)

        return render(request, 'myapp/payment.html', {
            'bus': bus,
            'seat_string': seat_string,
            'total_seats': total_seats,
            'total_cost': total_cost,
            'original_cost': total_cost,
            'dynamic_price': dynamic_price,
            'price_surge': dynamic_price > float(bus.base_price),
            'wallet_balance': float(profile.wallet_balance),
            'passenger_range': range(1, total_seats + 1),
        })

    return redirect('findbus')


# ── VALIDATE PROMO ────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def validate_promo(request):
    if request.method == 'POST':
        code  = request.POST.get('code', '').strip().upper()
        total = float(request.POST.get('total', 0))
        try:
            pc = PromoCode.objects.get(code__iexact=code)
            if pc.is_valid():
                disc = round(total * pc.discount_percent / 100, 2)
                return JsonResponse({'valid': True, 'discount': disc,
                                     'new_total': round(total - disc, 2),
                                     'percent': pc.discount_percent})
            return JsonResponse({'valid': False, 'msg': 'Code expired or limit reached.'})
        except PromoCode.DoesNotExist:
            return JsonResponse({'valid': False, 'msg': 'Invalid promo code.'})
    return JsonResponse({'valid': False, 'msg': 'Invalid request.'})


# ── MY BOOKINGS ───────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def seebookings(request):
    book_list = Book.objects.filter(userid=request.user.id)
    for b in book_list:
        if b.status == 'B' and b.date < date.today():
            b.status = 'X'; b.save()
    return render(request, 'myapp/booklist.html', {
        'book_list': book_list,
        'today': date.today(),
        'msg': None if book_list.exists() else 'No bookings yet.',
    })


# ── TRAVEL HISTORY ────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def travel_history(request):
    past = Book.objects.filter(userid=request.user.id, date__lt=date.today())
    return render(request, 'myapp/travel_history.html', {'past_bookings': past})


# ── CANCEL BOOKING ────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def cancellings(request):
    if request.method == 'POST':
        id_r = request.POST.get('bus_id', '').strip()
        try:
            book = Book.objects.get(id=id_r, userid=request.user.id)
        except Book.DoesNotExist:
            messages.error(request, "Booking not found or not yours.")
            return redirect('seebookings')

        if book.status == 'C':
            messages.error(request, "Already cancelled.")
            return redirect('seebookings')
        if book.date < date.today():
            messages.error(request, "Cannot cancel a past journey.")
            return redirect('seebookings')

        days_left = (book.date - date.today()).days
        refund_pct = 0.80 if days_left >= 2 else (0.50 if days_left == 1 else 0.0)
        refund_amt = round(float(book.price) * refund_pct, 2)

        with transaction.atomic():
            try:
                bus = Bus.objects.select_for_update().get(id=book.busid)
                bus.rem += book.nos; bus.save()
            except Bus.DoesNotExist:
                pass
            book.status = 'C'
            book.refund_amount = refund_amt
            book.refund_status = 'done' if refund_amt > 0 else 'none'
            book.save()
            if refund_amt > 0:
                profile = get_or_create_profile(request.user)
                profile.wallet_balance = float(profile.wallet_balance) + refund_amt
                profile.save()
                WalletTransaction.objects.create(
                    user=request.user, amount=refund_amt,
                    transaction_type='credit',
                    description=f'Refund for cancelled booking #{book.id}'
                )

        send_email_safe(
            'RouteWay – Booking Cancelled',
            f"Hello {request.user.username},\n\nBooking #{book.id} cancelled.\nRefund: Rs.{refund_amt} credited to wallet.\n\nRouteWay",
            request.user.email
        )
        messages.success(request, f"Booking cancelled. ₹{refund_amt} refunded to your wallet.")
    return redirect('seebookings')


# ── SUBMIT REVIEW ─────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def submit_review(request, book_id):
    book = get_object_or_404(Book, id=book_id, userid=request.user.id)
    if request.method == 'POST':
        rating  = int(request.POST.get('rating', 5))
        comment = request.POST.get('comment', '')
        try:
            bus = Bus.objects.get(id=book.busid)
            Review.objects.update_or_create(
                bus=bus, user=request.user,
                defaults={'booking': book, 'rating': rating, 'comment': comment}
            )
            messages.success(request, "Review submitted!")
        except Bus.DoesNotExist:
            messages.error(request, "Bus not found.")
        return redirect('travel_history')
    try:
        bus_obj = Bus.objects.get(id=book.busid)
    except Bus.DoesNotExist:
        bus_obj = None
    return render(request, 'myapp/review.html', {'book': book, 'bus': bus_obj})


# ── DOWNLOAD PDF TICKET ───────────────────────────────────────────────────────
@login_required(login_url='signin')
def download_ticket(request, book_id):
    book = get_object_or_404(Book, id=book_id, userid=request.user.id)
    if book.status == 'C':
        return render(request, 'myapp/error.html', {'error': 'Cancelled booking has no ticket.'})

    qr_data = (f"RouteWay Ticket\nID:{book.id}\nPassenger:{book.name}\n"
               f"Bus:{book.bus_name}\nFrom:{book.source}\nTo:{book.dest}\n"
               f"Seats:{book.seat_numbers}\nDate:{book.date}\nFare:Rs.{book.price}")
    qr = qrcode.make(qr_data)
    qr_buf = BytesIO(); qr.save(qr_buf); qr_buf.seek(0)

    pdf_buf = BytesIO()
    p = canvas.Canvas(pdf_buf, pagesize=A4)
    W, H = A4

    p.setFillColorRGB(0.04, 0.27, 0.62)
    p.rect(0, H-100, W, 100, fill=True, stroke=False)
    p.setFillColorRGB(1,1,1)
    p.setFont("Helvetica-Bold", 30)
    p.drawCentredString(W/2, H-50, "ROUTEWAY")
    p.setFont("Helvetica", 13)
    p.drawCentredString(W/2, H-72, "Bus Ticket  |  Boarding Pass")

    p.setStrokeColorRGB(0.8,0.8,0.8); p.setDash(4,3)
    p.line(40, H-110, W-40, H-110); p.setDash()

    p.setFillColorRGB(0.95,0.97,1)
    p.rect(40, H-165, W-80, 45, fill=True, stroke=False)
    p.setFillColorRGB(0.04,0.27,0.62)
    p.setFont("Helvetica-Bold", 18)
    p.drawString(60, H-137, book.source)
    p.drawRightString(W-60, H-137, book.dest)
    p.setFont("Helvetica", 14)
    p.drawCentredString(W/2, H-137, "─────────→")

    p.setFillColorRGB(0,0,0)
    fields = [
        ("Booking ID", f"# {book.id}"),
        ("Passenger", book.name),
        ("Bus", book.bus_name),
        ("Date", str(book.date)),
        ("Departure", str(book.time)),
        ("Seat(s)", book.seat_numbers or str(book.nos)),
        ("Total Fare", f"Rs. {book.price}"),
    ]
    if book.discount_amount:
        fields.append(("Discount", f"Rs. {book.discount_amount}"))
    y = H-195
    for label, val in fields:
        p.setFont("Helvetica-Bold", 11); p.setFillColorRGB(0.4,0.4,0.4)
        p.drawString(60, y, label + ":")
        p.setFont("Helvetica", 11); p.setFillColorRGB(0,0,0)
        p.drawString(210, y, str(val)); y -= 26

    if book.passenger_details:
        try:
            passengers = json.loads(book.passenger_details)
            y -= 5
            p.setFont("Helvetica-Bold", 11); p.setFillColorRGB(0.04,0.27,0.62)
            p.drawString(60, y, "Passengers:"); y -= 18
            for i, ps in enumerate(passengers, 1):
                p.setFont("Helvetica", 10); p.setFillColorRGB(0,0,0)
                p.drawString(70, y, f"{i}. {ps.get('name','')}  Age:{ps.get('age','')}  {ps.get('gender','')}"); y -= 16
        except Exception:
            pass

    qr_img = ImageReader(qr_buf)
    p.drawImage(qr_img, W-185, H-380, width=140, height=140)
    p.setFont("Helvetica", 8); p.setFillColorRGB(0.5,0.5,0.5)
    p.drawCentredString(W-115, H-388, "Scan to verify")

    p.setStrokeColorRGB(0.04,0.27,0.62); p.setLineWidth(1.5)
    p.line(40, 60, W-40, 60)
    p.setFont("Helvetica-Bold", 11); p.setFillColorRGB(0.04,0.27,0.62)
    p.drawCentredString(W/2, 42, "Thank you for travelling with RouteWay!")
    p.setFont("Helvetica", 9); p.setFillColorRGB(0.5,0.5,0.5)
    p.drawCentredString(W/2, 26, "Carry this ticket. Support: support@routeway.in")

    p.showPage(); p.save(); pdf_buf.seek(0)
    resp = HttpResponse(pdf_buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="RouteWay_Ticket_{book.id}.pdf"'
    return resp


# ── VERIFY TICKET ─────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def verify_ticket(request):
    result = None
    if request.method == 'POST':
        book_id = request.POST.get('book_id', '').strip()
        try:
            book = Book.objects.get(id=book_id)
            result = {'found': True, 'book': book}
        except Book.DoesNotExist:
            result = {'found': False}
    return render(request, 'myapp/verify_ticket.html', {'result': result})


# ── PROFILE ───────────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def profile(request):
    prof = get_or_create_profile(request.user)
    transactions = WalletTransaction.objects.filter(user=request.user).order_by('-created_at')[:10]
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_profile':
            request.user.email = request.POST.get('email', request.user.email)
            request.user.first_name = request.POST.get('first_name', '')
            request.user.last_name  = request.POST.get('last_name', '')
            request.user.save()
            prof.phone = request.POST.get('phone', '')
            if 'photo' in request.FILES:
                prof.photo = request.FILES['photo']
            prof.save()
            messages.success(request, "Profile updated!")
        elif action == 'change_password':
            old_pw = request.POST.get('old_password')
            new_pw = request.POST.get('new_password')
            if request.user.check_password(old_pw):
                if len(new_pw) >= 6:
                    request.user.set_password(new_pw)
                    request.user.save()
                    login(request, request.user)
                    messages.success(request, "Password changed!")
                else:
                    messages.error(request, "Password must be at least 6 characters.")
            else:
                messages.error(request, "Old password incorrect.")
        return redirect('profile')
    return render(request, 'myapp/profile.html', {'prof': prof, 'transactions': transactions})


# ── ADMIN DASHBOARD ───────────────────────────────────────────────────────────
@login_required(login_url='signin')
def admin_dashboard(request):
    if not request.user.is_staff:
        return redirect('home')
    today = date.today()
    total_revenue   = Book.objects.filter(status__in=['B','X']).aggregate(t=Sum('price'))['t'] or 0
    total_bookings  = Book.objects.filter(status__in=['B','X']).count()
    total_cancelled = Book.objects.filter(status='C').count()
    total_users     = User.objects.count()
    last7, labels = [], []
    for i in range(6,-1,-1):
        d = today - timedelta(days=i)
        last7.append(Book.objects.filter(booked_at__date=d).count())
        labels.append(d.strftime('%d %b'))
    top_routes = (Book.objects.filter(status__in=['B','X'])
                  .values('source','dest').annotate(cnt=Count('id')).order_by('-cnt')[:5])
    monthly_rev, monthly_labels = [], []
    for i in range(5,-1,-1):
        m = (today.replace(day=1) - timedelta(days=i*28))
        rev = Book.objects.filter(booked_at__year=m.year, booked_at__month=m.month,
                                  status__in=['B','X']).aggregate(t=Sum('price'))['t'] or 0
        monthly_rev.append(float(rev)); monthly_labels.append(m.strftime('%b %Y'))
    return render(request, 'myapp/admin_dashboard.html', {
        'total_revenue': total_revenue, 'total_bookings': total_bookings,
        'total_cancelled': total_cancelled, 'total_users': total_users,
        'last7': json.dumps(last7), 'labels': json.dumps(labels),
        'top_routes': top_routes,
        'monthly_rev': json.dumps(monthly_rev), 'monthly_labels': json.dumps(monthly_labels),
        'recent_bookings': Book.objects.order_by('-booked_at')[:10],
    })


# ── EXPORT EXCEL ──────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def export_bookings_excel(request):
    if not request.user.is_staff:
        return redirect('home')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Bookings"
    headers = ['ID','Passenger','Email','Bus','From','To','Seats','Qty','Price','Date','Status','Booked At']
    hf = PatternFill("solid", fgColor="0D3A8E")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hf; cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')
    for r, b in enumerate(Book.objects.all().order_by('-booked_at'), 2):
        ws.append([b.id, b.name, b.email, b.bus_name, b.source, b.dest,
                   b.seat_numbers, b.nos, float(b.price), str(b.date),
                   dict(Book.TICKET_STATUSES).get(b.status, b.status),
                   str(b.booked_at.strftime('%Y-%m-%d %H:%M') if b.booked_at else '')])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="RouteWay_Bookings.xlsx"'
    wb.save(resp); return resp


# ── MONTHLY REPORT ────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def monthly_report(request):
    if not request.user.is_staff:
        return redirect('home')
    today = date.today()
    mb = Book.objects.filter(booked_at__year=today.year, booked_at__month=today.month)
    total_rev  = mb.filter(status__in=['B','X']).aggregate(t=Sum('price'))['t'] or 0
    total_cnt  = mb.filter(status__in=['B','X']).count()
    cancel_cnt = mb.filter(status='C').count()
    pdf_buf = BytesIO(); p = canvas.Canvas(pdf_buf, pagesize=A4); W, H = A4
    p.setFillColorRGB(0.04,0.27,0.62); p.rect(0,H-90,W,90,fill=True,stroke=False)
    p.setFillColorRGB(1,1,1); p.setFont("Helvetica-Bold",24)
    p.drawCentredString(W/2, H-45, "ROUTEWAY – Monthly Report")
    p.setFont("Helvetica",12); p.drawCentredString(W/2, H-68, today.strftime('%B %Y'))
    p.setFillColorRGB(0,0,0); p.setFont("Helvetica",12)
    p.drawString(60, H-130, f"Total Bookings: {total_cnt}")
    p.drawString(60, H-150, f"Total Revenue: Rs. {total_rev}")
    p.drawString(60, H-170, f"Cancellations: {cancel_cnt}")
    p.drawString(60, H-190, f"Generated: {today}")
    y = H-230; p.setFont("Helvetica-Bold",11)
    for b in mb[:40]:
        if y < 60: p.showPage(); y = H-60
        p.setFont("Helvetica",9)
        p.drawString(60, y, f"#{b.id}  {b.name[:14]:<14}  {b.source}→{b.dest}  {b.date}  Rs.{b.price}  {b.status}")
        y -= 16
    p.showPage(); p.save(); pdf_buf.seek(0)
    resp = HttpResponse(pdf_buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="RouteWay_Report_{today.strftime("%b_%Y")}.pdf"'
    return resp


# ── OPERATOR REGISTER ─────────────────────────────────────────────────────────
def operator_register(request):
    if request.method == 'POST':
        name    = request.POST.get('name','').strip()
        email   = request.POST.get('email','').strip()
        pw      = request.POST.get('password','')
        company = request.POST.get('company','').strip()
        phone   = request.POST.get('phone','').strip()
        if User.objects.filter(username=name).exists():
            return render(request, 'myapp/operator_register.html', {'error': 'Username taken.'})
        user = User.objects.create_user(username=name, email=email, password=pw)
        Operator.objects.create(user=user, company_name=company, phone=phone)
        login(request, user)
        messages.success(request, "Operator account created! Awaiting admin approval.")
        return redirect('home')
    return render(request, 'myapp/operator_register.html')


# ── SIGNUP ────────────────────────────────────────────────────────────────────
def signup(request):
    context = {}
    if request.method == 'POST':
        name_r = request.POST.get('name','').strip()
        email_r = request.POST.get('email','').strip()
        password_r = request.POST.get('password','')
        if not name_r or not email_r or not password_r:
            context['error'] = "All fields are required."
        elif len(password_r) < 6:
            context['error'] = "Password must be at least 6 characters."
        elif User.objects.filter(username=name_r).exists():
            context['error'] = "Username already taken."
        elif User.objects.filter(email=email_r).exists():
            context['error'] = "Email already registered."
        else:
            user = User.objects.create_user(username=name_r, email=email_r, password=password_r)
            UserProfile.objects.get_or_create(user=user)
            login(request, user)
            return redirect('home')
    return render(request, 'myapp/signup.html', context)


# ── SIGNIN ────────────────────────────────────────────────────────────────────
def signin(request):
    if request.user.is_authenticated:
        return redirect('home')
    context = {}
    if request.method == 'POST':
        name_r = request.POST.get('name','').strip()
        pw_r   = request.POST.get('password','')
        if not name_r or not pw_r:
            context['error'] = "Please enter username and password."
        else:
            user = authenticate(request, username=name_r, password=pw_r)
            if user:
                login(request, user)
                if user.is_staff:
                    return redirect('admin_dashboard')
                return redirect('home')
            else:
                context['error'] = "Invalid username or password."
    return render(request, 'myapp/signin.html', context)


# ── SIGNOUT ───────────────────────────────────────────────────────────────────
def signout(request):
    logout(request); return redirect('signin')


# ── SEND OTP ──────────────────────────────────────────────────────────────────
def send_otp(request):
    otp = random.randint(1000, 9999)
    request.session['otp'] = otp
    messages.info(request, f"Demo OTP: {otp}")
    return redirect('signin')


# ── SUCCESS ───────────────────────────────────────────────────────────────────
@login_required(login_url='signin')
def success(request):
    return render(request, 'myapp/success.html', {'user': request.user})
